# coding=utf-8
from django.contrib.admin import site
from django.contrib.contenttypes.models import ContentType
from django.db import connection
from django.http.response import HttpResponseRedirect
from django.utils.encoding import force_text
from django.template.response import TemplateResponse
from django.contrib import messages
from django.utils.translation import ugettext_lazy as _
from django.contrib.contenttypes.models import ContentType
import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.utils.http import urlquote


def pay_action(request,model, object_id):
    title = "Are you sure?"
    ct = ContentType.objects.get(app_label='selfhelp',model=model)
    obj = ct.get_object_for_this_type(id=int(object_id))
    opts = obj._meta
    objects_name = force_text(opts.verbose_name)

    if model == 'reimbursement':
        loan = obj.loan
        amount = obj.logout_amount
        if loan and (amount is None or amount< 0):
            messages.error(request, u'您选择了借款单据，但是未正确填写核销金额，请在\'财务信息\'栏目中更正')
            return HttpResponseRedirect("/admin/selfhelp/%s/%s"%(model, object_id))

    if request.POST.get("post"):
        try:
            obj.action_pay(request)
            messages.success(request,_('action successfully'))
        except Exception,e:
            messages.error(request,e)

        return HttpResponseRedirect("/admin/selfhelp/%s/%s"%(model,object_id))

    context = dict(
        site.each_context(request),
        title=title,
        opts=opts,
        objects_name=objects_name,
        object=obj,
        action_name='pay'
    )
    request.current_app = site.name

    return TemplateResponse(request,'admin/invent/stockin/in_confirmation.html', context)


def export_excel(request, form):
    # 生成一个工作簿（Excel）
    wb = Workbook()
    wb.encoding = 'utf-8'
    # 获取第一个工作表（sheet1）
    sheet1 = wb.active
    sheet1.title = u'报销信息'
    row_one = ['单据编号', '创建人', '标题', '描述信息', '项目', '部门', '金额']
    for i in range(1, len(row_one)+1):
        # 从第一行开始写，列号也是从1开始
        sheet1.cell(row=1, column=i).value = row_one[i-1]
    # 获取到工作表的最大行数并加1
    max_row = sheet1.max_row + 1
    obj_info = [form.instance.code, form.instance.creator, form.instance.title, form.instance.description,
                form.instance.project.name, form.instance.org.name, float(form.instance.logout_amount)]
    # 将每一个对象的所有字段的信息写入一行内
    for x in range(1, len(obj_info)+1):
        sheet1.cell(row=max_row, column=x).value = obj_info[x-1]
    # 准备写入到IO中
    output = BytesIO()
    # 将Excel文件内容保存到IO中
    wb.save(output)
    # 重新定位到开始
    output.seek(0)
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # 给文件名中添加日期时间
    file_name = '报销信息%s.xls' % ctime
    # 使用urlquote()方法解决中文无法使用的问题
    file_name = urlquote(file_name)
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response
